import unittest
from yuanfang1987.commentFunctions.browsers import *
from openpyxl.workbook import workbook
from openpyxl.writer.excel import ExcelWriter


class MyTestCase(unittest.TestCase):
    # run firefox browser with a profile
    def setUp(self):
        self.dic = readBrowserSettings()
        self.driver = openBrowser(d=self.dic)
        # get the excel file that store test cases
        self.workbook = getworkbook()
        # get the excel file that store test data
        self.wbTestData = getTestDataworkbook()
        # create a new work book, in order to write result
        self.wb_result = workbook.Workbook()
        self.ew = ExcelWriter(workbook=self.wb_result)
        self.ws_result = self.wb_result.worksheets[0]
        self.ws_result.title = "Run Result"
        set_report_title(self.ws_result)

    def test_run_framework(self):
        # get the module that is marked as "Y" which means ready to run.
        self.modulesList = get_module_list(self.workbook,"Modules")
        for module in range(len(self.modulesList)):
            # get the test case sheet that is ready to run
            self.caseSheet = self.workbook.get_sheet_by_name(self.modulesList[module])
            self.rowNumber = self.caseSheet.get_highest_row()
            self.columnNumber = self.caseSheet.get_highest_column()
            self.pageList = list()
            self.actionList = list()
            iNext = 2
            for rowN in range(1,self.rowNumber+1):
                self.runResultdict = dict()
                self.steps = list()
                self.runResultdict["Continue"] = True
                self.runResultdict["Status"] = "Passed"
                self.runResultdict["Steps"] = self.steps
                self.runFlag = self.caseSheet.cell(row=rowN,column=1).value
                # get page list
                if self.runFlag == "Run":
                    self.pageList.clear()
                    for colN in range(4,self.columnNumber+1):
                        self.strValue = self.caseSheet.cell(row=rowN,column=colN).value
                        if self.strValue is not None:
                            self.pageList.append(self.strValue)
                # get action list
                elif self.runFlag == "Y":
                    # get the basic info of the current case and write them into report later.
                    self.caseid = self.caseSheet.cell(row=rowN,column=2).value
                    self.casename = self.caseSheet.cell(row=rowN,column=3).value
                    self.runResultdict["caseid"] = self.caseid
                    self.runResultdict["casename"] = self.casename
                    self.actionList.clear()
                    for colN in range(len(self.pageList)):
                        self.strValue = self.caseSheet.cell(row=rowN,column=colN+4).value
                        self.actionList.append(self.strValue)
                    # start to run case row by row
                    for i in range(len(self.pageList)):
                        self.page = self.pageList[i]
                        # create the page object
                        self.objPage = create_page_object(self.page,self.driver)
                        if self.actionList[i] is not None and self.runResultdict["Continue"]:
                            # get action name and test data
                            self.arrayTemp = self.actionList[i].split(";")
                            self.actionName = self.arrayTemp[0]
                            if len(self.arrayTemp) > 0:
                                self.dataNo = self.arrayTemp[1]
                                self.liTestData = getTestData(self.wbTestData,self.pageList[i],self.dataNo)
                            # run action on the current page.
                            self.objPage.run(self.actionName,self.liTestData,self.runResultdict)
                    #after finish one case, write result to report and reset the dict.
                    write_report(self.ws_result,iNext,self.runResultdict)
                    iNext = iNext + 1
                    #self.liTestData.clear(), seems no need to clear test data.
                    self.runResultdict.clear()

    def tearDown(self):
        self.driver.quit()
        self.ew.save("D:\\PythonWorkSpace\\yuanfang1987\\Data\\writeExcelTest0075.xlsx")

if __name__ == '__main__':
    unittest.main()
