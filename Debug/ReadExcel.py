from openpyxl.reader.excel import load_workbook

wb = load_workbook(filename="D:\\PythonWorkspace\\yuanfang\\Data\\TestCase.xlsx")

ws = wb.get_sheet_by_name("Modules")

print("hehe,github..")

rowNumber = ws.get_highest_row()
columnNumber = ws.get_highest_column()

for rowN in range(2,rowNumber+1):
    for colN in range(1,columnNumber+1):
        print("row: ",rowN,";col: ",colN,";value: ",ws.cell(row=rowN,column=colN).value)

