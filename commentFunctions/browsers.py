from selenium import webdriver
from openpyxl.reader.excel import load_workbook
from yuanfang1987.BusinessLogic.BaiduPage import Baidu


def openBrowser(d=dict()):
    dic = d
    strBrowserName = dic["BrowserName"]
    strURL = dic["URL"]
    if strBrowserName == "firefox":
        profileDir = "C:\\firefoxProfile"
        profile = webdriver.FirefoxProfile(profileDir)
        driver = webdriver.Firefox(profile)
        driver.get(strURL)
        driver.implicitly_wait(30)
    elif strBrowserName == "chrome":
        print("new a chrome browser here.")
    elif strBrowserName == "ie":
        print("new a ie browser here.")
    else:
        print("report error message that the browser is not defined.")
    return driver


def readBrowserSettings():
    d = dict()
    strFilePath = "D:\\PythonWorkSpace\\yuanfang1987\\Data\\Settings.xlsx"
    wb = load_workbook(strFilePath)
    ws = wb.get_sheet_by_name("browserSettings")
    for i in range(1,7):
        d[ws.cell(row=1,column=i).value] = ws.cell(row=2,column=i).value

    return d


def getworkbook():
    strFilePath = "D:\\PythonWorkSpace\\yuanfang1987\\Data\\TestCase.xlsx"
    wb = load_workbook(strFilePath)
    return wb


def get_module_list(wb,modulename):
    ws = wb.get_sheet_by_name(modulename)
    rowNumber = ws.get_highest_row()
    li = list()
    for rowN in range(1,rowNumber+1):
        if ws.cell(row=rowN,column=1).value == "Y":
            li.append(ws.cell(row=rowN,column=2).value)

    return li


def create_page_object(strPageName,gdriver):
    page = ""
    if strPageName == "Baidu":
        page = Baidu(gdriver)
    elif strPageName == "SearchResultPage":
        print("haha")
    return page


def getTestDataworkbook():
    strFilePath = "D:\\PythonWorkSpace\\yuanfang1987\\Data\\TestData.xlsx"
    wb = load_workbook(strFilePath)
    return wb


def getTestData(wb,sheetname,dataNo):
    ws = wb.get_sheet_by_name(sheetname)
    rowNumber = ws.get_highest_row()
    columnNumber = ws.get_highest_column()
    li = list()
    for rowN in range(1,rowNumber+1):
        if ws.cell(row=rowN,column=1).value == dataNo:
            di = dict()
            for colN in range(1,columnNumber+1):
                di[ws.cell(row=1,column=colN+1).value] = ws.cell(row=rowN,column=colN+1).value
            li.append(di)

    return li


def write_report(worksheet,intNext,dictResult):
    caseID = dictResult["caseid"]
    caseName = dictResult["casename"]
    status = dictResult["Status"]
    stepList = dictResult["Steps"]
    temp = "\n "
    steps = temp.join(stepList)
    #for i in range(len(stepList)):
    #   strSteps = strSteps + "Stpe[" + i + "]: " + stepList[i] + ";"
    # set the content
    worksheet.cell(row=intNext,column=1).value = status
    worksheet.cell(row=intNext,column=2).value = caseID
    worksheet.cell(row=intNext,column=3).value = caseName
    worksheet.cell(row=intNext,column=4).value = steps


def set_report_title(worksheet):
    worksheet.cell("A1").value = "Status"
    worksheet.cell("B1").value = "CaseID"
    worksheet.cell("C1").value = "Case Name"
    worksheet.cell("D1").value = "Test Stpes"
